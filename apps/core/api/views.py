# In apps/core/views/api.py

from rest_framework import status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import (
    ListAPIView, RetrieveAPIView, CreateAPIView,
    UpdateAPIView, DestroyAPIView, ListCreateAPIView,
    RetrieveUpdateDestroyAPIView
)
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.authentication import TokenAuthentication, SessionAuthentication
from rest_framework.throttling import UserRateThrottle
from rest_framework.pagination import PageNumberPagination
from rest_framework.filters import SearchFilter, OrderingFilter
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
import logging

from apps.core.permissions.mixins import (
    TenantRequiredMixin, RoleBasedViewMixin, PermissionRequiredMixin
)
from apps.core.services.audit_service import AuditService
from apps.core.utils.tenant import get_current_tenant
from apps.users.models import User

logger = logging.getLogger(__name__)


# ============================================================================
# CUSTOM PERMISSIONS
# ============================================================================

class TenantPermission(BasePermission):
    """Permission to ensure user belongs to current tenant"""
    
    def has_permission(self, request, view):
        tenant = getattr(request, 'tenant', None)
        if not tenant:
            return False
        
        # Check if user belongs to this tenant
        if hasattr(request.user, 'tenant'):
            return request.user.tenant == tenant
        
        return True


class RolePermission(BasePermission):
    """Permission based on user roles"""
    
    def has_permission(self, request, view):
        # Get required roles from view
        roles_required = getattr(view, 'roles_required', None)
        min_role_level = getattr(view, 'min_role_level', None)
        allow_superuser = getattr(view, 'allow_superuser', True)
        
        # Allow superuser if configured
        if allow_superuser and request.user.is_superuser:
            return True
        
        # Check specific roles
        if roles_required:
            if not isinstance(roles_required, (list, tuple)):
                roles_required = [roles_required]
            return request.user.role in roles_required
        
        # Check minimum role level
        if min_role_level:
            role_levels = User.ROLE_LEVELS
            user_level = role_levels.get(request.user.role, 0)
            required_level = role_levels.get(min_role_level, 0)
            return user_level >= required_level
        
        return True


class PermissionCombined(BasePermission):
    """Combine multiple permissions"""
    
    def __init__(self, permissions):
        self.permissions = permissions
    
    def has_permission(self, request, view):
        for permission in self.permissions:
            if not permission().has_permission(request, view):
                return False
        return True


# ============================================================================
# CUSTOM PAGINATION
# ============================================================================

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100
    
    def get_paginated_response(self, data):
        return Response({
            'count': self.page.paginator.count,
            'next': self.get_next_link(),
            'previous': self.get_previous_link(),
            'results': data,
            'page': self.page.number,
            'pages': self.page.paginator.num_pages,
            'page_size': self.get_page_size(self.request)
        })


# ============================================================================
# BASE API VIEWS
# ============================================================================

class BaseAPIView(APIView):
    """
    Base API view with authentication, tenant isolation, and audit logging
    """
    # Authentication
    authentication_classes = [TokenAuthentication, SessionAuthentication]
    permission_classes = [IsAuthenticated, TenantPermission, RolePermission]
    
    # Throttling
    throttle_classes = [UserRateThrottle]
    
    # Tenant configuration
    tenant_field = 'tenant'
    tenant_required = True
    
    # Role configuration
    roles_required = None
    min_role_level = None
    allow_superuser = True
    
    # Audit configuration
    audit_enabled = True
    audit_resource_type = None
    audit_action = None
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.tenant = None
    
    def initial(self, request, *args, **kwargs):
        """Initialize API view with tenant context"""
        super().initial(request, *args, **kwargs)
        
        # Set tenant from request
        self.tenant = getattr(request, 'tenant', None)
        
        # Validate tenant if required
        if self.tenant_required and not self.tenant:
            logger.warning(f"Tenant context missing in {self.__class__.__name__}")
    
    def get_audit_resource_type(self):
        """Get resource type for audit logging"""
        if self.audit_resource_type:
            return self.audit_resource_type
        
        # Try to determine from model
        if hasattr(self, 'model'):
            return self.model.__name__
        
        # Use class name as fallback
        return self.__class__.__name__.replace('APIView', '')
    
    def get_audit_action(self):
        """Get action for audit logging"""
        if self.audit_action:
            return self.audit_action
        
        # Map HTTP method to audit action
        method = self.request.method.upper()
        action_map = {
            'GET': 'READ',
            'POST': 'CREATE',
            'PUT': 'UPDATE',
            'PATCH': 'UPDATE',
            'DELETE': 'DELETE'
        }
        return action_map.get(method, 'READ')
    
    def get_queryset(self):
        """Get base queryset with tenant filtering"""
        if not hasattr(self, 'model'):
            raise NotImplementedError("View must define 'model' attribute")
        
        queryset = self.model.objects.all()
        
        # Apply tenant filtering
        if self.tenant and hasattr(self.model, 'tenant'):
            queryset = queryset.filter(tenant=self.tenant)
        
        # Apply active filter for soft-delete models
        try:
            self.model._meta.get_field('is_active')
            queryset = queryset.filter(is_active=True)
        except Exception:
            pass
        
        return queryset
    
    def perform_audit(self, instance=None, extra_data=None):
        """Perform audit logging"""
        if not self.audit_enabled:
            return
        
        try:
            AuditService.create_audit_entry(
                action=self.get_audit_action(),
                resource_type=self.get_audit_resource_type(),
                user=self.request.user,
                request=self.request,
                instance=instance,
                extra_data=extra_data or {}
            )
        except Exception as e:
            logger.error(f"Failed to create audit entry: {e}")
    
    def handle_exception(self, exc):
        """Handle exceptions with proper API response"""
        logger.error(f"API error in {self.__class__.__name__}: {exc}", exc_info=True)
        
        # Return appropriate error response
        if isinstance(exc, PermissionDenied):
            return Response(
                {'error': 'Permission denied', 'detail': str(exc)},
                status=status.HTTP_403_FORBIDDEN
            )
        elif isinstance(exc, Http404):
            return Response(
                {'error': 'Not found', 'detail': str(exc)},
                status=status.HTTP_404_NOT_FOUND
            )
        elif isinstance(exc, ValidationError):
            return Response(
                {'error': 'Validation error', 'detail': exc.message_dict},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Generic error
        return Response(
            {'error': 'Internal server error', 'detail': str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


class BaseListAPIView(BaseAPIView, ListAPIView):
    """
    Base list API view with filtering, searching, and pagination
    """
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, SearchFilter, OrderingFilter]
    
    # Filter configuration
    filterset_fields = []
    search_fields = []
    ordering_fields = ['created_at', 'updated_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Get filtered queryset"""
        queryset = super().get_queryset()
        
        # Apply additional filters
        queryset = self.apply_custom_filters(queryset)
        
        return queryset
    
    def apply_custom_filters(self, queryset):
        """Override to add custom filtering logic"""
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Override list to include audit logging"""
        response = super().list(request, *args, **kwargs)
        
        # Perform audit
        self.perform_audit(extra_data={
            'action': 'LIST',
            'count': response.data.get('count', 0),
            'filters': dict(request.query_params)
        })
        
        return response
    
    def get_serializer_context(self):
        """Add tenant context to serializer"""
        context = super().get_serializer_context()
        context['tenant'] = self.tenant
        context['request'] = self.request
        return context


class BaseRetrieveAPIView(BaseAPIView, RetrieveAPIView):
    """
    Base retrieve API view with object permission checking
    """
    object_permission_required = None
    
    def get_object(self):
        """Get object with permission checking"""
        obj = super().get_object()
        
        # Check object-level permissions
        if not self.has_object_permission(obj):
            raise PermissionDenied(
                f"You don't have permission to access this {self.model.__name__}"
            )
        
        return obj
    
    def has_object_permission(self, obj):
        """Check object-level permissions"""
        user = self.request.user
        
        if user.is_superuser:
            return True
        
        if self.object_permission_required:
            return user.has_perm(self.object_permission_required, obj)
        
        # Default: check if user is in the same tenant
        if hasattr(obj, 'tenant') and hasattr(user, 'tenant'):
            return obj.tenant == user.tenant
        
        # Check if user created the object
        if hasattr(obj, 'created_by'):
            return obj.created_by == user
        
        return True
    
    def retrieve(self, request, *args, **kwargs):
        """Override retrieve to include audit logging"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        
        # Perform audit
        self.perform_audit(instance=instance)
        
        return Response(serializer.data)


class BaseCreateAPIView(BaseAPIView, CreateAPIView):
    """
    Base create API view with tenant assignment
    """
    def perform_create(self, serializer):
        """Assign tenant and created_by to the instance"""
        # Assign tenant
        if self.tenant and 'tenant' not in serializer.validated_data:
            serializer.validated_data['tenant'] = self.tenant
        
        # Assign created_by
        if self.request.user.is_authenticated and 'created_by' not in serializer.validated_data:
            serializer.validated_data['created_by'] = self.request.user
        
        instance = serializer.save()
        
        # Perform audit
        self.perform_audit(
            instance=instance,
            extra_data={'created_via': 'api'}
        )
        
        return instance
    
    def create(self, request, *args, **kwargs):
        """Override create to include audit logging"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        instance = self.perform_create(serializer)
        
        headers = self.get_success_headers(serializer.data)
        return Response(
            serializer.data,
            status=status.HTTP_201_CREATED,
            headers=headers
        )


class BaseUpdateAPIView(BaseAPIView, UpdateAPIView):
    """
    Base update API view with change tracking
    """
    audit_changes = True
    
    def get_object(self):
        """Get object with permission checking"""
        obj = super().get_object()
        
        # Check object-level permissions
        if not self.has_object_permission(obj):
            raise PermissionDenied(
                f"You don't have permission to update this {self.model.__name__}"
            )
        
        return obj
    
    def perform_update(self, serializer):
        """Track changes and assign updated_by"""
        # Store old instance for change tracking
        old_instance = None
        if self.audit_changes:
            old_instance = self.model.objects.get(pk=serializer.instance.pk)
        
        # Assign updated_by
        if self.request.user.is_authenticated and 'updated_by' not in serializer.validated_data:
            serializer.validated_data['updated_by'] = self.request.user
        
        instance = serializer.save()
        
        # Perform audit with changes
        if self.audit_enabled and old_instance:
            try:
                from apps.core.services.audit_service import AuditService
                AuditService.log_update(
                    user=self.request.user,
                    instance=instance,
                    old_instance=old_instance,
                    request=self.request,
                    extra_data={'updated_via': 'api'}
                )
            except Exception as e:
                logger.error(f"Failed to log update: {e}")
        
        return instance


class BaseDestroyAPIView(BaseAPIView, DestroyAPIView):
    """
    Base destroy API view with soft delete support
    """
    soft_delete = True
    require_delete_reason = False
    
    def get_object(self):
        """Get object with permission checking"""
        obj = super().get_object()
        
        # Check object-level permissions
        if not self.has_object_permission(obj):
            raise PermissionDenied(
                f"You don't have permission to delete this {self.model.__name__}"
            )
        
        return obj
    
    def perform_destroy(self, instance):
        """Perform soft or hard delete"""
        # Check if model has is_active field for soft delete
        has_is_active = False
        try:
            instance._meta.get_field('is_active')
            has_is_active = True
        except Exception:
            pass
        
        if self.soft_delete and has_is_active:
            # Get deletion reason from request
            delete_reason = self.request.data.get('deletion_reason', '')
            delete_category = self.request.data.get('deletion_category', '')
            
            if self.require_delete_reason and (not delete_reason or not delete_category):
                raise ValidationError({
                    'deletion_reason': 'Deletion reason is required',
                    'deletion_category': 'Deletion category is required'
                })
            
            # Call soft delete method
            instance.delete(
                user=self.request.user,
                reason=delete_reason,
                category=delete_category
            )
            
            # Audit soft deletion
            self.perform_audit(
                instance=instance,
                extra_data={
                    'hard_delete': False,
                    'deletion_reason': delete_reason,
                    'deletion_category': delete_category
                }
            )
        else:
            # Hard delete
            instance.delete()
            
            # Audit hard deletion
            self.perform_audit(
                instance=instance,
                extra_data={'hard_delete': True}
            )


class BaseListCreateAPIView(BaseListAPIView, BaseCreateAPIView, ListCreateAPIView):
    """
    Combined list and create API view
    """
    pass


class BaseRetrieveUpdateDestroyAPIView(BaseRetrieveAPIView, BaseUpdateAPIView, 
                                       BaseDestroyAPIView, RetrieveUpdateDestroyAPIView):
    """
    Combined retrieve, update, and destroy API view
    """
    pass


# ============================================================================
# CUSTOM API VIEWS
# ============================================================================

class BulkOperationAPIView(BaseAPIView):
    """
    API view for bulk operations
    """
    def post(self, request, *args, **kwargs):
        """Handle bulk operations"""
        action = request.data.get('action')
        selected_ids = request.data.get('selected_ids', [])
        
        if not action or not selected_ids:
            return Response(
                {'error': 'Action and selected items required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                result = self.perform_bulk_action(action, selected_ids)
            
            # Perform audit
            self.perform_audit(extra_data={
                'action': f'BULK_{action.upper()}',
                'selected_ids': selected_ids,
                'count': len(selected_ids),
                'result': result
            })
            
            return Response({
                'success': True,
                'message': f'Bulk action completed successfully',
                'result': result
            })
            
        except Exception as e:
            logger.error(f"Bulk operation error: {e}", exc_info=True)
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def perform_bulk_action(self, action, selected_ids):
        """Override to implement bulk actions"""
        raise NotImplementedError("Subclasses must implement perform_bulk_action")


class ExportAPIView(BaseListAPIView):
    """
    API view for exporting data
    """
    export_formats = ['csv', 'excel', 'json']
    
    def get(self, request, *args, **kwargs):
        """Handle export request"""
        export_format = request.query_params.get('format', 'json')
        
        if export_format == 'csv':
            return self.export_csv()
        elif export_format == 'excel':
            return self.export_excel()
        elif export_format == 'json':
            return self.export_json()
        else:
            return Response(
                {'error': f'Unsupported export format: {export_format}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def export_csv(self):
        """Export data as CSV"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="export.csv"'
        
        writer = csv.writer(response)
        
        # Write headers
        headers = self.get_export_headers()
        writer.writerow(headers)
        
        # Write data
        queryset = self.get_queryset()
        for obj in queryset:
            writer.writerow(self.get_export_row(obj))
        
        return response
    
    def export_excel(self):
        """Export data as Excel"""
        import openpyxl
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Write headers
        headers = self.get_export_headers()
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        queryset = self.get_queryset()
        for row_idx, obj in enumerate(queryset, 2):
            row_data = self.get_export_row(obj)
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="export.xlsx"'
        wb.save(response)
        
        return response
    
    def export_json(self):
        """Export data as JSON"""
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        
        response = Response(serializer.data)
        response['Content-Disposition'] = 'attachment; filename="export.json"'
        return response
    
    def get_export_headers(self):
        """Override to provide export headers"""
        raise NotImplementedError("Subclasses must implement get_export_headers")
    
    def get_export_row(self, obj):
        """Override to provide export row data"""
        raise NotImplementedError("Subclasses must implement get_export_row")


class DashboardAPIView(BaseAPIView):
    """
    API view for dashboard statistics
    """
    def get(self, request, *args, **kwargs):
        """Get dashboard statistics"""
        tenant = get_current_tenant()
        
        stats = {
            'students': self.get_student_stats(tenant),
            'finance': self.get_finance_stats(tenant),
            'academics': self.get_academic_stats(tenant),
            'library': self.get_library_stats(tenant),
            'hostel': self.get_hostel_stats(tenant),
            'inventory': self.get_inventory_stats(tenant),
            'security': self.get_security_stats(tenant),
            'events': self.get_event_stats(tenant),
        }
        
        return Response(stats)
    
    def get_student_stats(self, tenant):
        """Get student statistics"""
        from apps.students.models import Student
        
        return {
            'total': Student.objects.filter(tenant=tenant).count(),
            'active': Student.objects.filter(tenant=tenant, status='ACTIVE').count(),
            'alumni': Student.objects.filter(tenant=tenant, status='ALUMNI').count(),
        }
    
    def get_finance_stats(self, tenant):
        """Get finance statistics"""
        from apps.finance.models import Invoice, Payment
        
        return {
            'total_invoices': Invoice.objects.filter(tenant=tenant).count(),
            'total_revenue': Payment.objects.filter(tenant=tenant).aggregate(
                total=Sum('amount')
            )['total'] or 0,
            'pending_invoices': Invoice.objects.filter(
                tenant=tenant, status='PENDING'
            ).count(),
        }
    
    def get_academic_stats(self, tenant):
        """Get academic statistics"""
        from apps.academics.models import SchoolClass, Course
        
        return {
            'total_classes': SchoolClass.objects.filter(tenant=tenant).count(),
            'total_courses': Course.objects.filter(tenant=tenant).count(),
        }
    
    def get_library_stats(self, tenant):
        """Get library statistics"""
        from apps.library.models import Book, BookIssue
        
        return {
            'total_books': Book.objects.filter(tenant=tenant).count(),
            'books_issued': BookIssue.objects.filter(
                tenant=tenant, return_date__isnull=True
            ).count(),
        }
    
    def get_hostel_stats(self, tenant):
        """Get hostel statistics"""
        from apps.hostel.models import Hostel, HostelAllocation
        
        return {
            'total_hostels': Hostel.objects.filter(tenant=tenant).count(),
            'total_allocations': HostelAllocation.objects.filter(
                tenant=tenant, is_active=True
            ).count(),
        }
    
    def get_inventory_stats(self, tenant):
        """Get inventory statistics"""
        from apps.inventory.models import Item
        
        return {
            'total_items': Item.objects.filter(tenant=tenant, is_active=True).count(),
            'low_stock': Item.objects.filter(
                tenant=tenant, is_active=True, current_stock__lte=10
            ).count(),
        }
    
    def get_security_stats(self, tenant):
        """Get security statistics"""
        from apps.security.models import SecurityIncident, AuditLog
        
        return {
            'open_incidents': SecurityIncident.objects.filter(
                tenant=tenant, status='OPEN'
            ).count(),
            'total_audits': AuditLog.objects.filter(tenant=tenant).count(),
        }
    
    def get_event_stats(self, tenant):
        """Get event statistics"""
        from apps.events.models import Event
        from django.utils import timezone
        
        return {
            'upcoming_events': Event.objects.filter(
                tenant=tenant, start_date__gte=timezone.now().date()
            ).count(),
        }


class GlobalSearchAPIView(BaseAPIView):
    """
    API view for global search
    """
    def get(self, request, *args, **kwargs):
        """Perform global search"""
        query = request.query_params.get('q', '').strip()
        
        if not query:
            return Response({'error': 'Search query required'}, status=400)
        
        results = {}
        
        # Search students (if user has permission)
        if request.user.role in ['ADMIN', 'SUPER_ADMIN', 'PRINCIPAL', 'VICE_PRINCIPAL', 'TEACHER']:
            results['students'] = self.search_students(query)
        
        # Search staff (if user has permission)
        if request.user.role in ['ADMIN', 'SUPER_ADMIN', 'PRINCIPAL', 'VICE_PRINCIPAL', 'HR']:
            results['staff'] = self.search_staff(query)
        
        # Search books (if user has permission)
        if request.user.role in ['ADMIN', 'SUPER_ADMIN', 'LIBRARIAN', 'TEACHER']:
            results['books'] = self.search_books(query)
        
        # Search invoices (if user has permission)
        if request.user.role in ['ADMIN', 'SUPER_ADMIN', 'ACCOUNTANT']:
            results['invoices'] = self.search_invoices(query)
        
        return Response(results)
    
    def search_students(self, query):
        """Search students"""
        from apps.students.models import Student
        
        students = Student.objects.filter(
            Q(first_name__icontains=query) | 
            Q(last_name__icontains=query) |
            Q(admission_number__icontains=query) |
            Q(email__icontains=query),
            tenant=self.tenant
        ).select_related('current_class', 'section')[:10]
        
        from apps.students.api.serializers import StudentListSerializer
        return StudentListSerializer(students, many=True).data
    
    def search_staff(self, query):
        """Search staff"""
        from apps.hr.models import Staff
        
        staff = Staff.objects.filter(
            Q(user__first_name__icontains=query) | 
            Q(user__last_name__icontains=query) |
            Q(employee_id__icontains=query) |
            Q(user__email__icontains=query),
            tenant=self.tenant
        ).select_related('user', 'department', 'designation')[:10]
        
        from apps.hr.api.serializers import StaffListSerializer
        return StaffListSerializer(staff, many=True).data
    
    def search_books(self, query):
        """Search books"""
        from apps.library.models import Book
        
        books = Book.objects.filter(
            Q(title__icontains=query) |
            Q(author__icontains=query) |
            Q(isbn__icontains=query),
            tenant=self.tenant
        )[:10]
        
        from apps.library.api.serializers import BookSerializer
        return BookSerializer(books, many=True).data
    
    def search_invoices(self, query):
        """Search invoices"""
        from apps.finance.models import Invoice
        
        invoices = Invoice.objects.filter(
            Q(invoice_number__icontains=query) |
            Q(student__first_name__icontains=query) |
            Q(student__last_name__icontains=query),
            tenant=self.tenant
        ).select_related('student')[:10]
        
        from apps.finance.api.serializers import InvoiceListSerializer
        return InvoiceListSerializer(invoices, many=True).data