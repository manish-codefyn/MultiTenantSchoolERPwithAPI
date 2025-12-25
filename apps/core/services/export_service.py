import csv
import json
from django.http import HttpResponse
from django.utils import timezone
import logging

logger = logging.getLogger(__name__)

class ExportService:
    """
    Service for handling data exports
    """
    
    @staticmethod
    def export_to_csv(queryset, filename, headers, row_callback):
        """
        Export queryset to CSV response
        """
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(headers)
        
        for obj in queryset:
            writer.writerow(row_callback(obj))
            
        return response

    @staticmethod
    def export_to_excel(queryset, filename, headers, row_callback):
        """
        Export queryset to Excel response
        """
        try:
            import openpyxl
        except ImportError:
            logger.error("openpyxl not installed")
            return HttpResponse("Excel export not available", status=501)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            
        # Write data
        for row_idx, obj in enumerate(queryset, 2):
            row_data = row_callback(obj)
            for col_idx, value in enumerate(row_data, 1):
                # Handle timezone aware datetimes for Excel
                if hasattr(value, 'tzinfo'):
                    value = value.replace(tzinfo=None)
                ws.cell(row=row_idx, column=col_idx, value=value)
                
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        wb.save(response)
        
        return response

    @staticmethod
    def export_to_json(queryset, filename, serializer_class, many=True):
        """
        Export queryset to JSON response
        """
        from rest_framework.response import Response
        serializer = serializer_class(queryset, many=many)
        response = Response(serializer.data)
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        return response
