import csv
import logging
from django.db import transaction
from django.core.files.uploadedfile import UploadedFile
import openpyxl

from apps.students.models import Student
from apps.students.services.student_service import StudentService
from apps.core.services.notification_service import NotificationService
from apps.academics.models import SchoolClass, Section, AcademicYear

logger = logging.getLogger(__name__)

class StudentImportService:
    """
    Service for importing students from files (CSV/Excel)
    """
    
    def __init__(self, tenant, created_by):
        self.tenant = tenant
        self.created_by = created_by
        self.results = {
            'created': 0,
            'updated': 0,
            'failed': 0,
            'errors': []
        }
        
    def import_from_file(self, file_obj: UploadedFile, create_user_accounts=False, send_welcome_emails=False):
        """
        Import students from uploaded file
        """
        try:
            if file_obj.name.endswith('.csv'):
                self._import_csv(file_obj, create_user_accounts, send_welcome_emails)
            elif file_obj.name.endswith(('.xlsx', '.xls')):
                self._import_excel(file_obj, create_user_accounts, send_welcome_emails)
            else:
                raise ValueError("Unsupported file format. Please use CSV or Excel.")
                
            return self.results
            
        except Exception as e:
            logger.error(f"Import error: {e}", exc_info=True)
            self.results['errors'].append(f"File processing error: {str(e)}")
            return self.results

    def _import_csv(self, file_obj, create_user, send_email):
        decoded_file = file_obj.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded_file)
        self._process_rows(reader, create_user, send_email)

    def _import_excel(self, file_obj, create_user, send_email):
        wb = openpyxl.load_workbook(file_obj)
        ws = wb.active
        
        # Get headers
        headers = [cell.value for cell in ws[1]]
        
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
            
        self._process_rows(rows, create_user, send_email)

    def _process_rows(self, rows, create_user, send_email):
        for index, row in enumerate(rows, start=1):
            try:
                with transaction.atomic():
                    self._process_single_row(row, index, create_user, send_email)
            except Exception as e:
                self.results['failed'] += 1
                self.results['errors'].append(f"Row {index}: {str(e)}")

    def _process_single_row(self, row, index, create_user, send_email):
        # Basic validation and data extraction
        first_name = row.get('First Name') or row.get('first_name')
        last_name = row.get('Last Name') or row.get('last_name')
        
        if not first_name or not last_name:
            raise ValueError("First Name and Last Name are required")

        # Get or create student
        # This is a simplified logic. In production, need robust matching.
        student_data = {
            'first_name': first_name,
            'last_name': last_name,
            'personal_email': row.get('Email') or row.get('email'),
            'mobile_primary': row.get('Phone') or row.get('phone'),
            'date_of_birth': row.get('DOB') or row.get('date_of_birth'),
            'gender': (row.get('Gender') or row.get('gender') or 'O')[0].upper(),
        }
        
        # Handle Class/Section
        class_name = row.get('Class') or row.get('class')
        section_name = row.get('Section') or row.get('section')
        
        if class_name:
            school_class = SchoolClass.objects.filter(
                tenant=self.tenant, name__iexact=class_name
            ).first()
            if school_class:
                student_data['current_class'] = school_class
        
        if section_name:
            section = Section.objects.filter(
                tenant=self.tenant, name__iexact=section_name, school_class=student_data.get('current_class')
            ).first()
            if section:
                student_data['section'] = section
                
        # Create student using service
        student, errors = StudentService.create_student(
            student_data, 
            self.created_by, 
            self.tenant
        )
        
        if student:
            self.results['created'] += 1
            
            if create_user:
                student.create_user_account()
                
            if send_email:
                NotificationService.send_student_registration_notification(
                    student, self.created_by, self.tenant
                )
        else:
            raise ValueError(f"Validation failed: {', '.join(errors)}")
